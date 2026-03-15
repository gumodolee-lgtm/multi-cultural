"""
내보내기 서비스 — DB 데이터를 Excel(.xlsx) 또는 CSV로 내보낸다.

openpyxl이 설치되어 있으면 Excel, 없으면 CSV로 폴백한다.
"""
from __future__ import annotations

import csv
import logging
from datetime import datetime
from pathlib import Path

from app.services.data_provider import DataProvider

logger = logging.getLogger(__name__)


class ExportService:
    """데이터 내보내기 서비스"""

    @staticmethod
    def export_all(output_dir: str = "data/export") -> str:
        """뉴스, 법령, 지원사업을 각 시트(또는 파일)로 내보낸다.

        Returns:
            저장된 파일 경로
        """
        out_dir = Path(output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        try:
            return ExportService._export_xlsx(out_dir, timestamp)
        except ImportError:
            logger.info("openpyxl 미설치 — CSV로 내보내기")
            return ExportService._export_csv(out_dir, timestamp)

    @staticmethod
    def _export_xlsx(out_dir: Path, timestamp: str) -> str:
        """openpyxl을 사용하여 Excel 파일로 내보낸다."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # -- 뉴스 시트 --
        ws_news = wb.active
        ws_news.title = "뉴스"
        news_headers = ["제목", "출처", "카테고리", "발행일", "URL", "북마크"]
        _write_header(ws_news, news_headers)

        for item in DataProvider.get_all_news():
            ws_news.append([
                item["title"],
                item["source"],
                item["category"],
                item["published"],
                item["url"],
                "Y" if item["is_bookmarked"] else "",
            ])

        # -- 법령 시트 --
        ws_laws = wb.create_sheet("법령")
        law_headers = ["법령명", "법령코드", "분류", "개정일", "시행일", "AI요약", "북마크"]
        _write_header(ws_laws, law_headers)

        for item in DataProvider.get_all_laws():
            ws_laws.append([
                item["name"],
                item["law_code"],
                item["category"],
                item["amended_date"],
                item["effective_date"],
                item["summary"],
                "Y" if item["is_bookmarked"] else "",
            ])

        # -- 지원사업 시트 --
        ws_support = wb.create_sheet("지원사업")
        support_headers = ["사업명", "주관기관", "기관유형", "지역", "대상", "혜택", "신청시작", "신청마감", "연락처", "URL", "북마크"]
        _write_header(ws_support, support_headers)

        for item in DataProvider.get_all_support():
            ws_support.append([
                item["name"],
                item["organizer"],
                item["org_type"],
                item["region"],
                item["target_group"],
                item["benefit"],
                item["apply_start"],
                item["apply_end"],
                item["contact"],
                item["url"],
                "Y" if item["is_bookmarked"] else "",
            ])

        # 열 너비 자동 조정
        for ws in [ws_news, ws_laws, ws_support]:
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        filename = f"michub_export_{timestamp}.xlsx"
        path = out_dir / filename
        wb.save(str(path))
        logger.info("Excel 내보내기 완료: %s", path)
        return str(path)

    @staticmethod
    def _export_csv(out_dir: Path, timestamp: str) -> str:
        """CSV 파일 3개로 내보낸다."""
        # 뉴스
        news_path = out_dir / f"news_{timestamp}.csv"
        with open(news_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["제목", "출처", "카테고리", "발행일", "URL", "북마크"])
            for item in DataProvider.get_all_news():
                writer.writerow([
                    item["title"], item["source"], item["category"],
                    item["published"], item["url"],
                    "Y" if item["is_bookmarked"] else "",
                ])

        # 법령
        laws_path = out_dir / f"laws_{timestamp}.csv"
        with open(laws_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["법령명", "법령코드", "분류", "개정일", "시행일", "AI요약", "북마크"])
            for item in DataProvider.get_all_laws():
                writer.writerow([
                    item["name"], item["law_code"], item["category"],
                    item["amended_date"], item["effective_date"], item["summary"],
                    "Y" if item["is_bookmarked"] else "",
                ])

        # 지원사업
        support_path = out_dir / f"support_{timestamp}.csv"
        with open(support_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["사업명", "주관기관", "기관유형", "지역", "대상", "혜택", "신청시작", "신청마감", "연락처", "URL", "북마크"])
            for item in DataProvider.get_all_support():
                writer.writerow([
                    item["name"], item["organizer"], item["org_type"],
                    item["region"], item["target_group"], item["benefit"],
                    item["apply_start"], item["apply_end"],
                    item["contact"], item["url"],
                    "Y" if item["is_bookmarked"] else "",
                ])

        logger.info("CSV 내보내기 완료: %s", out_dir)
        return str(out_dir)


def _write_header(ws, headers: list[str]) -> None:
    """시트 헤더를 작성한다 (스타일 포함)."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    except ImportError:
        ws.append(headers)
